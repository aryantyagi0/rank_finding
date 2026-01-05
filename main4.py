import pandas as pd
import requests
import time
import os
import re
from urllib.parse import urlparse, parse_qs, unquote
# Extracts real URLs from Google redirect links()
from typing import TypedDict, Optional
from dotenv import load_dotenv
from openpyxl import load_workbook
from langchain_openai import ChatOpenAI
from langgraph.graph import StateGraph, END
import webbrowser
# Opens Google Search or Maps automatically
from urllib.parse import quote
import subprocess
# Opens Chrome or Edge in Incognito mode
import shutil 
# Detects if Chrome or Edge exists on system
import sys


# =============================
# LOAD ENV
# =============================
load_dotenv()

SERP_API_KEY = os.getenv("SERP_API_KEY")
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")

if not SERP_API_KEY or not OPENAI_API_KEY:
    raise ValueError("❌ Missing API keys")

# =============================
# CONFIG
# =============================
INPUT_FILE = "input.xlsx"
SHEET_NAME = "Keywords"
LOCATION = "Noida, Uttar Pradesh, India"
MAPS_LL = "@28.5355,77.3910,14z"
RATE_LIMIT_DELAY = 2
OPEN_BROWSER = False   # set False for headless / automation runs
USE_INCOGNITO = False     # turn off if needed
BROWSER_DELAY = 2       # seconds between browser actions

# =============================
# LLM (AGENT BRAIN)
# =============================
llm = ChatOpenAI(
    model="gpt-4o-mini",
    temperature=0,
    api_key=OPENAI_API_KEY
)

# =============================
# DOMAIN HELPERS
# =============================
def normalize_domain(domain: str) -> str:
    if not isinstance(domain, str):
        return ""
    domain = domain.lower().strip()
    domain = re.sub(r'^https?://', '', domain)
    domain = re.sub(r'^www\.', '', domain)
    return domain.split('/')[0]
# #| Step              | Example                                       |
# | ----------------- | --------------------------------------------- |
# | lowercase         | `HTTPS://EXAMPLE.COM` → `https://example.com` |
# | remove https | `https://example.com` → `example.com`         |
# | remove www        | `www.example.com` → `example.com`             |
# | remove path       | `example.com/page` → `example.com`            |

def extract_domain(url: str) -> str:
    if not isinstance(url, str):
        return ""

    if "google.com/url" in url:
        try:
            qs = parse_qs(urlparse(url).query)
            url = qs.get("q", [""])[0]
            url = unquote(url)
        except:
            pass

    if not url.startswith("http"):
        url = "https://" + url

    try:
        return normalize_domain(urlparse(url).netloc)
    except:
        return ""
# Google search results often give links like:
# https://www.google.com/url?q=https://example.com/page
# This function:
# Finds the real link inside it
# Removes extra parts
# Returns only:example.com
def domain_matches(d1: str, d2: str) -> bool:
    d1, d2 = normalize_domain(d1), normalize_domain(d2)
    return (
        d1 == d2 or
        d1.endswith("." + d2) or
        d2.endswith("." + d1)
    )

def normalize_name(name: str) -> str:
    if not name:
        return ""
    name = name.lower()
    name = re.sub(r'[^a-z0-9 ]', '', name)  # remove symbols
    name = re.sub(r'\s+', ' ', name).strip()
    return name

# checking serp result domain against target domain

# =============================
# SERP SAFE REQUEST
# =============================
def safe_serp_request(params):
    for _ in range(3):
        try:
            r = requests.get("https://serpapi.com/search", params=params, timeout=40)
            r.raise_for_status()
            return r.json()
        except:
            time.sleep(2)
    return {}
# It calls the SERP API
# If it works → returns the data
# If it fails → waits 2 seconds and tries again
# It tries maximum 3 times
# If all fail → returns empty data {}

def open_incognito(url):
    """
    Try to open URL in Chrome Incognito or Edge InPrivate.
    Falls back to normal browser if unavailable.
    """
    try:
        # 1️ Try Chrome
        chrome_path = shutil.which("chrome") or shutil.which("chrome.exe")
        if chrome_path:
            subprocess.Popen([chrome_path, "--incognito", url])
            print("🕶️ Opened in Chrome Incognito")
            return

        # 2 Try Edge
        edge_path = shutil.which("msedge") or shutil.which("msedge.exe")
        if edge_path:
            subprocess.Popen([edge_path, "--inprivate", url])
            print("🕶️ Opened in Edge InPrivate")
            return

    except Exception as e:
        print(f"⚠️ Incognito failed: {e}")

    # 3️ Graceful fallback
    print("🌐 Fallback: opening in normal browser")
    webbrowser.open_new_tab(url)

# =============================
# TOOLS
# =============================
def get_organic_rank(keyword, domain):
    best_rank = None
    best_url = ""

    for page in range(0, 5):  # top 50
        start = page * 10
        params = {
            "engine": "google",
            "q": keyword,
            "location": LOCATION,
            "google_domain": "google.co.in",
            "hl": "en",
            "gl": "in",
            "start": start,
            "num": 10,
            "api_key": SERP_API_KEY
        }

        data = safe_serp_request(params)

        for result in data.get("organic_results", []):
            link = result.get("link", "")
            if domain_matches(extract_domain(link), domain):
                rank = start + result.get("position", 0)
                if best_rank is None or rank < best_rank:
                    best_rank = rank
                    best_url = link

        # stop early only if we already found best possible on this page
        if best_rank is not None:
            break

        time.sleep(RATE_LIMIT_DELAY)

    return (str(best_rank), best_url) if best_rank else ("Not Found", "")
def get_pack_rank(keyword: str, domain: str, business_name: str) -> Optional[str]:
    """
    Hybrid Local Pack Rank:
    1) Google Maps
    2) Fallback to Local Finder (tbm=lcl)
    """

    # ---------- 1️⃣ GOOGLE MAPS ----------
    params_maps = {
        "engine": "google_maps",
        "q": keyword,
        "ll": MAPS_LL,
        "api_key": SERP_API_KEY
    }

    data = safe_serp_request(params_maps)
    places = data.get("local_results", [])

    norm_business = normalize_name(business_name)

    for idx, place in enumerate(places, start=1):
        website = place.get("website", "")
        title = place.get("title", "")

        if website and domain_matches(website, domain):
            return str(idx)

        if norm_business and norm_business in normalize_name(title):
            return str(idx)

    # ---------- 2️⃣ LOCAL FINDER FALLBACK ----------
    params_lcl = {
        "engine": "google",
        "q": keyword,
        "tbm": "lcl",
        "location": LOCATION,
        "hl": "en",
        "gl": "in",
        "api_key": SERP_API_KEY
    }

    data = safe_serp_request(params_lcl)
    places = data.get("local_results", [])

    for idx, place in enumerate(places, start=1):
        website = place.get("website", "")
        title = place.get("title", "")

        if website and domain_matches(website, domain):
            return str(idx)

        if norm_business and norm_business in normalize_name(title):
            return str(idx)

    return "Not in pack"









def should_open_browser(state):
    """
    Agent decision:
    Open browser only for high-value results.
    """
    try:
        org_rank = int(state["organic_rank"])
    except:
        org_rank = None

    try:
        map_rank = int(state["maps_rank"])
    except:
        map_rank = None

    # Agent decision rules
    return (
        (org_rank is not None and org_rank <= 10) or
        (map_rank is not None and map_rank <= 5)
    )

def is_organic_found(state):
    try:
        return int(state["organic_rank"]) > 0
    except:
        return False


def is_maps_found(state):
    try:
        return int(state["maps_rank"]) > 0
    except:
        return False

def open_google_search(keyword):
    url = f"https://www.google.com/search?q={quote(keyword)}"
    if USE_INCOGNITO:
        open_incognito(url)
    else:
        webbrowser.open_new_tab(url)



def open_google_maps(keyword):
    url = f"https://www.google.com/maps/search/{quote(keyword)}"
    if USE_INCOGNITO:
        open_incognito(url)
    else:
        webbrowser.open_new_tab(url)



# =============================
# AGENT STATE
# =============================
class RankState(TypedDict):
    keyword: str
    domain: str
    organic_rank: Optional[str]
    organic_url: Optional[str]
    maps_rank: Optional[str]   # ← this is NOW pack_rank
    reasoning: Optional[str]
    business_name: str


# =============================
# AGENT NODES
# =============================
def organic_node(state: RankState):
    rank, url = get_organic_rank(state["keyword"], state["domain"])
    state["organic_rank"] = rank
    state["organic_url"] = url
    return state

def maps_node(state: RankState):
    state["maps_rank"] = get_pack_rank(
        state["keyword"],
        state["domain"],
        state["business_name"]
    )
    return state


def reasoning_node(state: RankState):
    prompt = f"""
You are an SEO reporting assistant.

STRICT RULES (must follow):
- Use ONLY the data provided below.
- Do NOT assume trends, causes, or improvements.
- Do NOT compare with previous periods.
- Do NOT add recommendations.
- Do NOT add external knowledge.
- Do NOT change numeric values.
- Do NOT speculate.
- Keep the explanation factual and neutral.

DATA:
Keyword: {state['keyword']}
Organic Rank (Google Search): {state['organic_rank']}
Local Pack Rank (Maps → Local Finder): {state['maps_rank']}


TASK:
Write 2–3 sentences explaining what these results indicate.
Mention that minor differences from the Google UI are expected.
"""

    state["reasoning"] = llm.invoke(prompt).content.strip()
    return state


# =============================
# AGENT GRAPH
# =============================
graph = StateGraph(RankState)

graph.add_node("organic", organic_node)
graph.add_node("maps", maps_node)
graph.add_node("reasoning", reasoning_node)

graph.set_entry_point("organic")
graph.add_edge("organic", "maps")
graph.add_edge("maps", "reasoning")
graph.add_edge("reasoning", END)

agent = graph.compile()

# =============================
# MAIN
# =============================

def auto_adjust_row_height(ws, row, col, text, col_width=55):
    """
    Adjust row height based on text length for wrapped cells.
    """
    if not text:
        return

    # Rough estimate: characters per line
    chars_per_line = col_width * 1.2
    lines = max(1, int(len(text) / chars_per_line) + 1)

    ws.row_dimensions[row].height = lines * 18


def main():
    wb = load_workbook(INPUT_FILE)
    ws = wb[SHEET_NAME]

    df_raw = pd.read_excel(INPUT_FILE, sheet_name=SHEET_NAME, header=None)

    header_idx = None
    for i in range(5):
        if "Local Keyword Ideas" in df_raw.iloc[i].astype(str).values:
            header_idx = i
            break

    if header_idx is None:
        raise ValueError("Header row with 'Local Keyword Ideas' not found")

    header_row = header_idx + 1

    GOOGLE_PLACES_COL = 34   # AH
    GOOGLE_LINKS_COL  = 35   # AI
    AGENT_REASON_COL  = 39   # AK

    # ---- Header ----
    from openpyxl.styles import Alignment

    ws.cell(row=header_row, column=AGENT_REASON_COL).value = "Agent Reasoning"
    ws.column_dimensions["AM"].width = 55

    header_cell = ws.cell(row=header_row, column=AGENT_REASON_COL)
    header_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[header_row].height = 28

    # ---- Prepare dataframe ----
    df = df_raw.copy()
    df.columns = df.iloc[header_idx]
    df = df.iloc[header_idx + 1:].reset_index(drop=False)
    df.rename(columns={"index": "excel_row"}, inplace=True)
    df["excel_row"] += 1

    df = df.dropna(subset=["Local Keyword Ideas", "Targeted Page"])
    df["Domain"] = df["Targeted Page"].apply(extract_domain)

    # ---- Process rows ----
    for _, row in df.iterrows():
        excel_row = int(row["excel_row"])
        keyword = row["Local Keyword Ideas"]

        if (
            ws.cell(row=excel_row, column=GOOGLE_LINKS_COL).value
            and ws.cell(row=excel_row, column=AGENT_REASON_COL).value
        ):
            print(f"⏭️ Skipping: {keyword}")
            continue

        print(f"🤖 Agent processing: {keyword}")

        state = agent.invoke({
            "keyword": keyword,
            "domain": row["Domain"],
            "organic_rank": None,
            "business_name": "OMKITCHEN",
            "organic_url": None,
            "maps_rank": None,
            "reasoning": None
        })

        # ---- Browser logic (SAFE) ----
        if OPEN_BROWSER and should_open_browser(state):

            organic_found = is_organic_found(state)
            maps_found = is_maps_found(state)

            if organic_found:
                print("👀 Opening Google Search")
                open_google_search(keyword)
                time.sleep(BROWSER_DELAY)

            if maps_found:
                print("📍 Opening Google Maps")
                open_google_maps(keyword)
                time.sleep(BROWSER_DELAY)

        # ---- Write to Excel ----
        ws.cell(row=excel_row, column=GOOGLE_PLACES_COL).value = state["maps_rank"]
        ws.cell(row=excel_row, column=GOOGLE_LINKS_COL).value = state["organic_rank"]

        cell = ws.cell(row=excel_row, column=AGENT_REASON_COL)
        cell.value = state["reasoning"]
        cell.alignment = Alignment(wrap_text=True, vertical="top")

        auto_adjust_row_height(
            ws,
            row=excel_row,
            col=AGENT_REASON_COL,
            text=state["reasoning"],
            col_width=55
        )

        time.sleep(RATE_LIMIT_DELAY)

    wb.save(INPUT_FILE)
    print("✅ Rankings + Agent Reasoning written into SAME file")

if __name__ == "__main__": 
    main()
